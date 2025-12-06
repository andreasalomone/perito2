import base64
import functools
import logging
import mimetypes
import os
import io
import pathlib
from typing import Any, Callable, Dict, List, TypeVar, Union, cast

import fitz  # PyMuPDF
import mailparser
import openpyxl
from PIL import Image
import re
import uuid
import zipfile
import xml.etree.ElementTree as ET

def sanitize_filename(filename: str) -> str:
    """
    Sanitizes a filename to prevent path traversal and remove dangerous characters.
    Keeps alphanumeric, dots, dashes, and underscores.
    """
    # Remove path components
    filename = os.path.basename(filename)
    # Replace anything that isn't alphanumeric, ., -, or _
    filename = re.sub(r'[^a-zA-Z0-9._-]', '_', filename)
    return filename



logger = logging.getLogger(__name__)

F = TypeVar("F", bound=Callable[..., Any])

def handle_extraction_errors(
    default_return_value: Any = None,
) -> Callable[[F], F]:
    """Decorator to handle common exceptions during file processing."""
    def decorator(func: F) -> F:
        @functools.wraps(func)
        def wrapper(file_path: str, *args: Any, **kwargs: Any) -> Any:
            try:
                return func(file_path, *args, **kwargs)
            except (ValueError, fitz.FileDataError, openpyxl.utils.exceptions.InvalidFileException) as e:
                # Domain errors: Return error dict for user visibility
                logger.warning(f"Domain error processing {file_path}: {e}")
                return [{
                    "type": "error",
                    "filename": os.path.basename(file_path),
                    "message": str(e),
                }]
            except MemoryError as e:
                # CRITICAL: OOM should propagate to monitoring systems
                logger.critical(f"CRITICAL: Out of memory processing {file_path}: {e}")
                raise
            except OSError as e:
                # CRITICAL: Disk full or I/O errors should propagate
                # ENOSPC (28) = No space left on device
                # EDQUOT (122) = Disk quota exceeded
                if hasattr(e, 'errno') and e.errno in [28, 122]:
                    logger.critical(f"CRITICAL: Disk full processing {file_path}: {e}")
                    raise
                # Other OS errors might be recoverable (permissions, etc) but often indicate system issues.
                # For safety in this refactor, we will RAISE them to be caught by the caller
                # unless we are sure they are user errors.
                logger.error(f"OS error processing {file_path}: {e}", exc_info=True)
                raise
            except Exception as e:
                # Unexpected system errors (Bug, etc): RAISE them.
                logger.error(f"Critical failure processing {file_path}: {e}", exc_info=True)
                raise
        return wrapper # type: ignore
    return decorator

# --- Standardized Extractors (All return List[Dict]) ---

@handle_extraction_errors()
def prepare_pdf_for_llm(pdf_path: str) -> List[Dict[str, Any]]:
    parts = []
    
    # 1. Vision Part (Always included)
    parts.append({
        "type": "vision",
        "path": pdf_path,
        "mime_type": "application/pdf",
        "filename": sanitize_filename(os.path.basename(pdf_path)),
    })

    # 2. Text Part (Optional, if text exists)
    try:
        doc = fitz.open(pdf_path)
        text_content = "".join(page.get_text() for page in doc)
        doc.close()
        
        # Improved check: Just check if there is any non-whitespace text
        if text_content.strip():
            parts.append({
                "type": "text",
                "content": text_content,
                "filename": f"{sanitize_filename(os.path.basename(pdf_path))} (extracted text)",
            })
    except Exception as e:
        logger.warning(f"Could not extract text from PDF: {e}")
        # Return an error object so the failure is visible in the extracted data
        parts.append({
            "type": "error",
            "filename": f"{sanitize_filename(os.path.basename(pdf_path))} (text extraction)",
            "message": f"Text extraction failed: {e}",
        })

    return parts

@handle_extraction_errors()
def prepare_image_for_llm(image_path: str) -> List[Dict[str, Any]]:
    # Validate image
    with Image.open(image_path) as img:
        img.verify()
        
    mime_type, _ = mimetypes.guess_type(image_path)
    if not mime_type or not mime_type.startswith("image/"):
        mime_type = "application/octet-stream"

    return [{
        "type": "vision",
        "path": image_path,
        "mime_type": mime_type,
        "filename": sanitize_filename(os.path.basename(image_path)),
    }]

@handle_extraction_errors()
def extract_text_from_docx(docx_path: str) -> List[Dict[str, Any]]:
    """
    Extracts text from DOCX using streaming XML parsing (SAX-like).
    This avoids loading the entire document object model into memory, 
    fixing the 'Major' memory vulnerability identified in the audit.
    """
    text_content = []
    try:
        with zipfile.ZipFile(docx_path) as zf:
            if "word/document.xml" not in zf.namelist():
                raise ValueError("Invalid DOCX: Missing word/document.xml")
                
            with zf.open("word/document.xml") as f:
                # iterparse is a streaming parser
                context = ET.iterparse(f, events=("end",))
                for event, elem in context:
                    # w:t is the tag for text in Word XML
                    if elem.tag.endswith("}t"):
                        if elem.text:
                            text_content.append(elem.text)
                    # Clean up element to free memory
                    elem.clear()
        
        full_text = "\n".join(text_content)
        
    except Exception as e:
        logger.warning(f"Fast extraction failed: {e}. Falling back to standard method if needed, or failing.")
        raise e

    return [{
        "type": "text",
        "content": full_text,
        "filename": sanitize_filename(os.path.basename(docx_path)),
    }]

@handle_extraction_errors()
def extract_text_from_xlsx(xlsx_path: str) -> List[Dict[str, Any]]:
    text_content = ""
    try:
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            text_content += f"--- Sheet: {sheet_name} ---\n"
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_values = [str(cell) if cell is not None else "" for cell in row]
                text_content += ",".join(row_values) + "\n"
            text_content += "\n"
        workbook.close()
    except Exception as e:
        logger.warning(f"Failed to read XLSX file {xlsx_path}: {e}")
        return [{
            "type": "error",
            "filename": sanitize_filename(os.path.basename(xlsx_path)),
            "message": f"Could not read Excel file content: {str(e)}",
        }]

    return [{
        "type": "text",
        "content": text_content,
        "filename": sanitize_filename(os.path.basename(xlsx_path)),
    }]

@handle_extraction_errors()
def extract_text_from_txt(txt_path: str) -> List[Dict[str, Any]]:
    # Prevent OOM on maliciously large text files
    MAX_TXT_SIZE = 10 * 1024 * 1024  # 10MB
    file_size = os.path.getsize(txt_path)
    if file_size > MAX_TXT_SIZE:
        logger.warning(f"Text file {txt_path} is too large ({file_size} bytes). Skipping.")
        return [{
            "type": "error",
            "filename": sanitize_filename(os.path.basename(txt_path)),
            "message": f"File too large ({file_size:,} bytes, max {MAX_TXT_SIZE:,} bytes)",
        }]
    
    with open(txt_path, "r", encoding="utf-8") as f:
        content = f.read()
    return [{
        "type": "text", 
        "content": content, 
        "filename": sanitize_filename(os.path.basename(txt_path))
    }]

@handle_extraction_errors()
def process_eml_file(eml_path: str, upload_folder: str, depth: int = 0) -> List[Dict[str, Any]]:
    """
    Processes an .eml file, extracting its text body and saving/processing attachments.
    Returns a FLAT LIST of dictionaries.
    """
    if depth > 3:
        logger.warning(f"Max recursion depth reached for {eml_path}")
        return [{
            "type": "error",
            "filename": os.path.basename(eml_path),
            "message": "Max recursion depth reached (nested attachments)",
        }]

    mail = mailparser.parse_from_file(eml_path)
    all_parts: List[Dict[str, Any]] = []

    # Extract plain text body
    text_content = ""
    if mail.text_plain:
        text_content = "\n".join(mail.text_plain)
    elif mail.text_html:
        text_content = "\n".join(mail.text_html)
    else:
        text_content = mail.body if mail.body else ""

    all_parts.append({
        "type": "text",
        "content": text_content,
        "filename": f"{sanitize_filename(os.path.basename(eml_path))} (body)",
        "original_filetype": "eml",
    })

    # Use StorageProvider to save attachments
    # storage = get_storage_provider() # Removed for now as we need local processing for extraction

    for attachment in mail.attachments:
        original_filename = attachment.get("filename", "untitled_attachment")

        # Check for excluded extensions
        _, ext = os.path.splitext(original_filename)
        ext = ext.lower()
        if ext in [".gif", ".mp4", ".avi", ".mov", ".mkv", ".webm"]:
            logger.info(
                f"Skipping excluded attachment type '{ext}' for file '{original_filename}' in {eml_path}"
            )
            continue

        payload = attachment.get("payload", "")

        if not payload:
            continue

        try:
            if isinstance(payload, str):
                payload = payload.strip()
                missing_padding = len(payload) % 4
                if missing_padding:
                    payload += "=" * (4 - missing_padding)

            decoded_payload = base64.b64decode(payload)
            
            # --- TINY IMAGE FILTER (Signatures/Icons) ---
            # 1. Check File Size (< 5KB)
            if len(decoded_payload) < 5 * 1024:
                # Potential signature/icon. Check dimensions if it's an image.
                if ext in [".png", ".jpg", ".jpeg", ".gif", ".webp"]:
                    try:
                        # We need to peek at dimensions without saving if possible, 
                        # or just save and check. Saving is safer/easier since we have the bytes.
                        # But let's try to be efficient and check bytes if possible, 
                        # or just rely on size for now? 
                        # User asked for "Option 1" which implies size AND dimensions.
                        
                        # Let's use PIL to check dimensions from bytes
                        with Image.open(io.BytesIO(decoded_payload)) as img:
                            width, height = img.size
                            if width < 100 and height < 100:
                                logger.info(f"Skipping tiny image attachment '{original_filename}' ({width}x{height}, {len(decoded_payload)} bytes) in {eml_path}")
                                continue
                    except Exception as e:
                        # If we can't parse it as an image, just proceed (safe failure)
                        logger.warning(f"Could not check dimensions of small image {original_filename}: {e}")
            # --------------------------------------------
            
            # Use the robust sanitize_filename utility
            safe_filename = sanitize_filename(original_filename)
            if not safe_filename:
                safe_filename = "attachment.bin"

            # FIX: Prevent filename collisions in recursive processing
            # Prepend a short UUID to ensure uniqueness in the flat upload folder
            unique_prefix = uuid.uuid4().hex[:8]
            safe_filename = f"{unique_prefix}_{safe_filename}"

            # Save to local temp folder for recursive processing
            # This ensures we can read it back with fitz/openpyxl/etc.
            local_path = os.path.join(upload_folder, safe_filename)
            
            with open(local_path, "wb") as f:
                f.write(decoded_payload)
            
            # Recurse
            if os.path.exists(local_path):
                attachment_parts = process_uploaded_file(local_path, upload_folder, depth=depth + 1)
                if isinstance(attachment_parts, list):
                    all_parts.extend(attachment_parts)
                else:
                    all_parts.append(attachment_parts)
            
            # Optional: If we wanted to persist to GCS, we would do it here.
            # But for report generation, extracting the content is the primary goal.

        except Exception as e:
            logger.error(f"Error processing attachment {original_filename}: {e}")

    return all_parts

def process_uploaded_file(filepath: str, upload_folder: str, depth: int = 0) -> List[Dict[str, Any]]:
    _, ext = os.path.splitext(filepath)
    ext = ext.lower()
    
    # Simple Dispatcher
    processors = {
        ".pdf": prepare_pdf_for_llm,
        ".png": prepare_image_for_llm,
        ".jpg": prepare_image_for_llm,
        ".jpeg": prepare_image_for_llm,
        ".webp": prepare_image_for_llm,
        ".gif": prepare_image_for_llm,
        ".docx": extract_text_from_docx,
        ".xlsx": extract_text_from_xlsx,
        ".txt": extract_text_from_txt,
    }
    
    if ext == ".eml":
        return process_eml_file(filepath, upload_folder, depth=depth)
    
    processor = processors.get(ext)
    
    if processor:
        result = processor(filepath)
        # Verify it is a list (POLA enforcement)
        if isinstance(result, dict): 
            return [result]
        return result
        
    return [{
        "type": "unsupported",
        "filename": os.path.basename(filepath),
        "message": f"Unsupported file type: {ext}"
    }]
