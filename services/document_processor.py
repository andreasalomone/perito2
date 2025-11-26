import base64  # Added for decoding attachments
import functools
import logging
import mimetypes  # Added for MIME type guessing
import os
import pathlib  # Added for path manipulation
from typing import Any, Callable, Dict, List, TypeVar, Union, cast  # Added cast

import fitz  # PyMuPDF for PDFs - still used for basic validation or if needed for non-AI tasks
import mailparser  # Added for .eml files
import openpyxl  # For XLSX files
from docx import Document as DocxDocument

# import pytesseract # For OCR - No longer needed for PDF/image text extraction by LLM
from PIL import (  # For image handling - potentially for validation, no longer for OCR
    Image,
)

logger = logging.getLogger(__name__)

# pytesseract.pytesseract.tesseract_cmd = r'<full_path_to_your_tesseract_executable>' # No longer needed

F = TypeVar("F", bound=Callable[..., Any])


def handle_extraction_errors(
    default_return_value: Any = None,
) -> Callable[[F], F]:  # Modified default
    """Decorator to handle common exceptions during file processing."""

    def decorator(func: F) -> F:
        @functools.wraps(func)
        def wrapper(file_path: str, *args: Any, **kwargs: Any) -> Any:
            try:
                return func(file_path, *args, **kwargs)
            except FileNotFoundError:
                logger.error(f"File not found: {file_path}")
                return {
                    "type": "error",
                    "filename": os.path.basename(file_path),
                    "message": "File not found",
                }
            # TesseractError no longer primary concern for LLM path
            # except pytesseract.TesseractError as te:
            #     logger.error(f"Tesseract OCR error for {file_path}: {te}", exc_info=True)
            #     return default_return_value
            except Image.UnidentifiedImageError:
                logger.error(f"Cannot identify image file: {file_path}", exc_info=True)
                return {
                    "type": "error",
                    "filename": os.path.basename(file_path),
                    "message": "Cannot identify image file",
                }
            except openpyxl.utils.exceptions.InvalidFileException as ofe:
                logger.error(
                    f"Invalid Excel (openpyxl) file format for {file_path}: {ofe}",
                    exc_info=True,
                )
                return {
                    "type": "error",
                    "filename": os.path.basename(file_path),
                    "message": f"Invalid Excel file: {ofe}",
                }
            except Exception as e:
                logger.error(
                    f"Unexpected error processing {file_path} with {func.__name__}: {e}",
                    exc_info=True,
                )
                return {
                    "type": "error",
                    "filename": os.path.basename(file_path),
                    "message": f"Unexpected error: {e}",
                }

        return wrapper  # type: ignore

    return decorator


# PDF and Image "extraction" functions will now just return file info for the LLM
@handle_extraction_errors()
def prepare_pdf_for_llm(pdf_path: str) -> Union[Dict[str, Any], List[Dict[str, Any]]]:
    """Prepares PDF file information for LLM processing, including text extraction if available."""
    # Basic validation: can it be opened?
    try:
        doc = fitz.open(pdf_path)
    except Exception as e:
        logger.error(f"PDF file {pdf_path} is likely corrupted or not a valid PDF: {e}")
        raise  # Re-raise to be caught by decorator

    vision_part = {
        "type": "vision",
        "path": pdf_path,
        "mime_type": "application/pdf",
        "filename": os.path.basename(pdf_path),
    }

    # Attempt text extraction
    text_content = ""
    try:
        for page in doc:
            text_content += page.get_text()
    except Exception as e:
        logger.warning(f"Failed to extract text from PDF {pdf_path}: {e}")
    finally:
        doc.close()

    # If we found significant text, return both vision and text parts
    # Threshold of 50 chars to avoid "scanned" PDFs that might have just a tiny bit of noise/metadata text
    if len(text_content.strip()) > 50:
        text_part = {
            "type": "text",
            "content": text_content,
            "filename": f"{os.path.basename(pdf_path)} (extracted text)",
        }
        return [vision_part, text_part]

    return vision_part


@handle_extraction_errors()
def prepare_image_for_llm(image_path: str) -> Dict[str, str]:
    """Prepares image file information for LLM processing."""
    # Basic validation: can it be opened by PIL?
    try:
        img = Image.open(image_path)
        img.close()
    except Exception as e:
        logger.error(
            f"Image file {image_path} is likely corrupted or not a valid image: {e}"
        )
        raise  # Re-raise

    mime_type, _ = mimetypes.guess_type(image_path)
    if not mime_type or not mime_type.startswith("image/"):
        logger.warning(
            f"Could not determine a valid image MIME type for {image_path}, defaulting to application/octet-stream."
        )
        mime_type = "application/octet-stream"  # Fallback, though Gemini might reject
    return {
        "type": "vision",
        "path": image_path,
        "mime_type": mime_type,
        "filename": os.path.basename(image_path),
    }


@handle_extraction_errors(
    {"type": "text", "content": "", "filename": ""}
)  # filename will be overwritten
def extract_text_from_docx(docx_path: str) -> Dict[str, str]:
    """Extracts text from a DOCX file."""
    doc = DocxDocument(docx_path)
    full_text: List[str] = [para.text for para in doc.paragraphs]
    return {
        "type": "text",
        "content": "\n".join(full_text),
        "filename": os.path.basename(docx_path),
    }


@handle_extraction_errors(
    {"type": "text", "content": "", "filename": ""}
)  # filename will be overwritten
def extract_text_from_xlsx(xlsx_path: str) -> Dict[str, str]:
    """Extracts text from an XLSX file, converting sheets to CSV-like text."""
    text_content: str = ""
    try:
        # Use read_only=True and data_only=True for better performance and stability with complex files
        workbook = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        for sheet_name in workbook.sheetnames:
            text_content += f"--- Sheet: {sheet_name} ---\n"
            sheet = workbook[sheet_name]
            for row in sheet.iter_rows(values_only=True):
                row_values: List[str] = [
                    str(cell) if cell is not None else "" for cell in row
                ]
                text_content += ",".join(row_values) + "\n"
            text_content += "\n"
        workbook.close()
    except Exception as e:
        logger.warning(f"Failed to read XLSX file {xlsx_path}: {e}")
        return {
            "type": "text",
            "content": f"[ERROR: Could not read Excel file content: {str(e)}]",
            "filename": os.path.basename(xlsx_path),
        }

    return {
        "type": "text",
        "content": text_content,
        "filename": os.path.basename(xlsx_path),
    }


@handle_extraction_errors(
    {"type": "text", "content": "", "filename": ""}
)  # filename will be overwritten
def extract_text_from_txt(txt_path: str) -> Dict[str, str]:
    """Extracts text from a TXT file."""
    with open(txt_path, "r", encoding="utf-8") as f:
        content = f.read()
    return {"type": "text", "content": content, "filename": os.path.basename(txt_path)}


@handle_extraction_errors(
    []
)  # On error, return an empty list to match the new signature
def process_eml_file(eml_path: str, upload_folder: str) -> List[Dict[str, Any]]:
    """
    Processes an .eml file, extracting its text body and saving/processing attachments.
    Returns a FLAT LIST of dictionaries, one for the text content and one for each processed attachment part.
    """
    mail = mailparser.parse_from_file(eml_path)

    # The list that will be returned, containing all content parts from the EML.
    all_parts: List[Dict[str, Any]] = []

    # Extract plain text body
    # Prioritize text_plain, then html (converted to text by mailparser), then body
    text_content = ""
    if mail.text_plain:
        text_content = "\n".join(mail.text_plain)
    elif mail.text_html:
        text_content = "\n".join(
            mail.text_html
        )  # mailparser often provides a text version from html
    else:
        text_content = mail.body if mail.body else ""
        if not text_content:
            logger.warning(
                f"EML file {eml_path} has no discernible text body (text_plain, text_html, or body)."
            )

    # Add the main email body as the first part
    all_parts.append(
        {
            "type": "text",
            "content": text_content,
            "filename": f"{os.path.basename(eml_path)} (body)",  # Clarify this is the body
            "original_filetype": "eml",
        }
    )

    # Ensure the 'out' directory for attachments exists within the upload_folder
    attachments_output_dir = pathlib.Path(upload_folder) / "email_attachments"
    attachments_output_dir.mkdir(parents=True, exist_ok=True)

    for attachment in mail.attachments:
        original_filename = attachment.get("filename", "untitled_attachment")
        payload = attachment.get("payload", "")

        if not payload:
            logger.warning(
                f"Attachment '{original_filename}' in {eml_path} has no payload. Skipping."
            )
            continue

        # CORRECTED LOGIC: Attempt to process ALL attachments, letting process_uploaded_file handle filtering.
        try:
            # Ensure payload is a string and strip whitespace
            if isinstance(payload, str):
                payload = payload.strip()
                # Fix incorrect padding
                missing_padding = len(payload) % 4
                if missing_padding:
                    payload += "=" * (4 - missing_padding)
            
            # Decode base64
            decoded_payload = base64.b64decode(payload)
            # Sanitize filename (simple sanitization)
            safe_filename = "".join(
                c if c.isalnum() or c in (".", "_", "-") else "_"
                for c in original_filename
            )
            if not safe_filename:
                content_type = attachment.get("mail_content_type", "")
                safe_filename = f"attachment_{content_type.split('/')[-1] if '/' in content_type else 'bin'}"

            attachment_save_path = attachments_output_dir / safe_filename

            # Avoid overwriting if file with same name exists by appending a number
            counter = 1
            temp_path = attachment_save_path
            while temp_path.exists():
                name_part, ext_part = temp_path.stem, temp_path.suffix
                # Remove previous counter if exists to avoid names like file_1_2_3
                if name_part.endswith(f"_{counter-1}") and counter > 1:
                    name_part = name_part[: -(len(str(counter - 1)) + 1)]
                temp_path = attachments_output_dir / f"{name_part}_{counter}{ext_part}"
                counter += 1
            attachment_save_path = temp_path

            with open(attachment_save_path, "wb") as f:
                f.write(decoded_payload)
            logger.info(
                f"Saved attachment '{attachment_save_path.name}' from {eml_path} to {attachments_output_dir}"
            )

            # Recursively process the saved attachment.
            # This will now correctly handle .docx, .xlsx, .txt, etc., in addition to PDF/images.
            attachment_parts = process_uploaded_file(
                str(attachment_save_path), upload_folder
            )

            if not attachment_parts:
                continue

            if isinstance(attachment_parts, list):
                # If it's a list (from a nested .eml), extend our list with its parts
                all_parts.extend(attachment_parts)
            elif attachment_parts.get("type") not in ["error", "unsupported"]:
                # If it's a dict for a valid file type, append it
                all_parts.append(cast(Dict[str, Any], attachment_parts))
            else:
                # Log error/unsupported cases but don't add them to the parts list
                logger.warning(
                    f"Could not process or skipped attachment '{attachment_save_path.name}' from {eml_path}. Info: {attachment_parts}"
                )

        except base64.binascii.Error as b64e:
            logger.error(
                f"Base64 decoding error for attachment '{original_filename}' in {eml_path}: {b64e}"
            )
        except IOError as ioe:
            logger.error(
                f"IOError saving attachment '{original_filename}' from {eml_path}: {ioe}"
            )
        except Exception as e:
            logger.error(
                f"Unexpected error processing attachment '{original_filename}' from {eml_path}: {e}",
                exc_info=True,
            )

    return all_parts


def process_uploaded_file(
    filepath: str, upload_folder: str
) -> Union[Dict[str, Any], List[Dict[str, Any]]]:
    """Determines file type and calls the appropriate processing function.

    Returns a dictionary for most file types, but a list of dictionaries for .eml files.
    """
    _, ext = os.path.splitext(filepath)
    ext = ext.lower()
    filename: str = os.path.basename(filepath)  # Moved here for all return paths

    processing_function: Union[Callable[..., Any], None] = None

    if ext == ".pdf":
        processing_function = prepare_pdf_for_llm
    elif ext in [".png", ".jpg", ".jpeg", ".webp", ".gif"]:  # Added more image types
        processing_function = prepare_image_for_llm
    elif ext == ".docx":
        processing_function = extract_text_from_docx
    elif ext == ".xlsx":
        processing_function = extract_text_from_xlsx
    elif ext == ".txt":
        processing_function = extract_text_from_txt
    elif ext == ".eml":
        # For .eml, we need to pass the upload_folder to save attachments
        # The functools.partial allows us to pre-fill the 'upload_folder' argument
        # This will return a List[Dict[str, Any]]
        processing_function = functools.partial(
            process_eml_file, upload_folder=upload_folder
        )
    else:
        logger.warning(f"Unsupported file type: {ext} for file {filepath}")
        return {
            "type": "unsupported",
            "filename": filename,
            "message": f"Unsupported file type: {ext}",
        }

    if processing_function:
        result = processing_function(filepath)

        # If the result is a dictionary (most cases), ensure filename is present.
        # If it's a list (from an .eml), its elements are assumed to be correctly formatted.
        if isinstance(result, dict):
            if "filename" not in result or not result["filename"]:
                result["filename"] = filename
        return result

    # Should not be reached if logic is correct, but as a fallback:
    logger.error(
        f"No processing function determined for {filepath}, though extension seemed supported."
    )
    return {
        "type": "error",
        "filename": filename,
        "message": "Internal error: No processing function found.",
    }
